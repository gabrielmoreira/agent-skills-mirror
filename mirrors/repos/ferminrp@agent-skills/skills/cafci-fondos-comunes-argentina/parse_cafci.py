#!/usr/bin/env python3
"""Normalize the CAFCI /pb_get daily XLSX into JSON for jq queries.

Usage:
    python3 parse_cafci.py <path-to-xlsx>

Writes JSON to stdout with shape:
    {
      "fecha_reporte": "YYYY-MM-DD",
      "categorias": ["Renta Variable Peso Argentina", ...],
      "fondos": [{nombre, categoria, moneda, region, horizonte, fecha,
                  vcp_actual, vcp_anterior, variacion_dia_pct, vcp_reexp_pesos,
                  variacion_mes_pct, variacion_ytd_pct, variacion_12m_pct,
                  cantidad_cuotapartes, patrimonio, market_share,
                  depositaria, codigo_cnv}, ...]
    }
"""
import json
import sys
from datetime import date, datetime

import openpyxl


def to_iso_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, (date, datetime)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None


def to_num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def to_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def parse(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    fondos = []
    current_cat = None
    fecha_reporte = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 9:
            continue
        nonempty = [(j, v) for j, v in enumerate(row) if v is not None and v != ""]
        if not nonempty:
            continue
        # Category row: only column 0 populated, no fecha, no vcp.
        if len(nonempty) == 1 and nonempty[0][0] == 0:
            cat = to_str(nonempty[0][1])
            if cat:
                current_cat = cat
            continue
        nombre = to_str(row[0])
        if not nombre:
            continue
        fecha = to_iso_date(row[4])
        if fecha and fecha_reporte is None:
            fecha_reporte = fecha
        fondos.append({
            "nombre": nombre,
            "categoria": current_cat,
            "moneda": to_str(row[1]),
            "region": to_str(row[2]),
            "horizonte": to_str(row[3]),
            "fecha": fecha,
            "vcp_actual": to_num(row[5]),
            "vcp_anterior": to_num(row[6]),
            "variacion_dia_pct": to_num(row[7]),
            "vcp_reexp_pesos": to_num(row[8]),
            "variacion_mes_pct": to_num(row[9]),
            "variacion_ytd_pct": to_num(row[10]),
            "variacion_12m_pct": to_num(row[11]),
            "cantidad_cuotapartes": to_num(row[12]),
            "patrimonio": to_num(row[14]),
            "market_share": to_num(row[16]),
            "depositaria": to_str(row[17]),
            "codigo_cnv": to_str(row[18]),
        })

    seen = set()
    categorias = []
    for f in fondos:
        c = f["categoria"]
        if c and c not in seen:
            seen.add(c)
            categorias.append(c)

    return {
        "fecha_reporte": fecha_reporte,
        "categorias": categorias,
        "fondos": fondos,
    }


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("usage: parse_cafci.py <path-to-xlsx>", file=sys.stderr)
        sys.exit(1)
    result = parse(sys.argv[1])
    json.dump(result, sys.stdout, ensure_ascii=False, separators=(",", ":"))
    sys.stdout.write("\n")
