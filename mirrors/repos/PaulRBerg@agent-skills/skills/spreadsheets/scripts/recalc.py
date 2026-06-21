#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.12"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Recalculate formulas in an .xlsx/.xlsm with headless LibreOffice, then audit it.

Saves the workbook in place and prints a JSON report whose `status` is one of:
success | errors_found | recalc_incomplete | error.

By default, every non-success status exits nonzero. Use --soft to preserve the
old report-only behavior for formula errors and incomplete recalculation.

Usage: uv run scripts/recalc.py <workbook> [timeout-seconds] [--soft]
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import sys
import tempfile
from collections import defaultdict
from pathlib import Path
from typing import Any, NoReturn

EXCEL_ERRORS = {"#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A", "#SPILL!", "#CALC!"}
LIBREOFFICE_ERROR = re.compile(r"Err:\d{3}")
MAX_CELLS_PER_ERROR = 20
DEFAULT_TIMEOUT = 60

MACRO_SUB = "RecalcAndSaveClose"
MACRO_XML = f"""<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
Sub {MACRO_SUB}()
    ThisComponent.calculateAll()
    ThisComponent.store()
    ThisComponent.close(True)
End Sub
</script:module>
"""
SCRIPT_XLC = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE library:libraries PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "libraries.dtd">
<library:libraries xmlns:library="http://openoffice.org/2000/library" xmlns:xlink="http://www.w3.org/1999/xlink">
 <library:library library:name="Standard" library:link="false"/>
</library:libraries>
"""
DIALOG_XLC = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE library:libraries PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "libraries.dtd">
<library:libraries xmlns:library="http://openoffice.org/2000/library" xmlns:xlink="http://www.w3.org/1999/xlink"/>
"""
SCRIPT_XLB = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE library:library PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "library.dtd">
<library:library xmlns:library="http://openoffice.org/2000/library" library:name="Standard" library:readonly="false" library:passwordprotected="false">
 <library:element library:name="Module1"/>
</library:library>
"""
DIALOG_XLB = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE library:library PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "library.dtd">
<library:library xmlns:library="http://openoffice.org/2000/library" library:name="Standard" library:readonly="false" library:passwordprotected="false"/>
"""


def report(payload: dict[str, Any], code: int = 0) -> NoReturn:
    print(json.dumps(payload, indent=2))
    sys.exit(code)


def find_soffice() -> str | None:
    on_path = shutil.which("soffice")
    if on_path:
        return on_path
    for candidate in (
        Path("/Applications/LibreOffice.app/Contents/MacOS/soffice"),
        Path.home() / "Applications/LibreOffice.app/Contents/MacOS/soffice",
    ):
        if candidate.is_file():
            return str(candidate)
    return None


def run_soffice(args: list[str], timeout: int) -> subprocess.CompletedProcess[str]:
    return subprocess.run(args, capture_output=True, text=True, timeout=timeout, check=False)


def profile_arg(profile_root: Path) -> str:
    return f"-env:UserInstallation={profile_root.resolve().as_uri()}"


def write_profile_macro(profile_root: Path) -> None:
    basic = profile_root / "user/basic"
    standard = basic / "Standard"
    standard.mkdir(parents=True, exist_ok=True)
    (basic / "script.xlc").write_text(SCRIPT_XLC, encoding="utf-8")
    (basic / "dialog.xlc").write_text(DIALOG_XLC, encoding="utf-8")
    (standard / "script.xlb").write_text(SCRIPT_XLB, encoding="utf-8")
    (standard / "dialog.xlb").write_text(DIALOG_XLB, encoding="utf-8")
    (standard / "Module1.xba").write_text(MACRO_XML, encoding="utf-8")


def ensure_macro(soffice: str, profile_root: Path) -> None:
    profile_root.mkdir(parents=True, exist_ok=True)
    run_soffice(
        [soffice, profile_arg(profile_root), "--headless", "--norestore", "--terminate_after_init"],
        timeout=60,
    )
    write_profile_macro(profile_root)


def recalculate(soffice: str, profile_root: Path, workbook: Path, timeout: int) -> subprocess.CompletedProcess[str]:
    uri = f"vnd.sun.star.script:Standard.Module1.{MACRO_SUB}?language=Basic&location=application"
    return run_soffice(
        [
            soffice,
            profile_arg(profile_root),
            "--headless",
            "--norestore",
            "--nolockcheck",
            uri,
            str(workbook),
        ],
        timeout=timeout,
    )


def audit(workbook: Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    with_values = load_workbook(workbook, data_only=True)
    try:
        with_formulas = load_workbook(workbook, data_only=False)
    except Exception:
        with_values.close()
        raise

    try:
        errors: dict[str, list[str]] = defaultdict(list)
        total_errors = 0
        total_formulas = 0
        uncached = 0

        for name in with_formulas.sheetnames:
            formula_ws = with_formulas[name]
            value_ws = with_values[name]
            for row in formula_ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        total_formulas += 1
                        if value_ws[cell.coordinate].value is None:
                            uncached += 1
            for row in value_ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        token = cell.value.strip()
                        if token in EXCEL_ERRORS or LIBREOFFICE_ERROR.fullmatch(token):
                            errors[token].append(f"{name}!{cell.coordinate}")
                            total_errors += 1
    finally:
        with_values.close()
        with_formulas.close()

    if total_errors:
        status = "errors_found"
    elif total_formulas and uncached == total_formulas:
        status = "recalc_incomplete"
    else:
        status = "success"

    return {
        "file": str(workbook),
        "status": status,
        "total_formulas": total_formulas,
        "uncached_formulas": uncached,
        "total_errors": total_errors,
        "errors": {
            kind: {"count": len(cells), "cells": cells[:MAX_CELLS_PER_ERROR]}
            for kind, cells in sorted(errors.items())
        },
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("timeout_seconds", nargs="?", type=int, default=DEFAULT_TIMEOUT)
    parser.add_argument("--soft", action="store_true", help="exit 0 after formula errors; still fails operational errors")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook = args.workbook.resolve()
    timeout = args.timeout_seconds
    if timeout < 1:
        report({"status": "error", "error": "timeout_seconds must be >= 1"}, code=1)
    if not workbook.is_file():
        report({"status": "error", "error": f"{workbook} does not exist"}, code=1)

    soffice = find_soffice()
    if soffice is None:
        report(
            {
                "status": "error",
                "error": "LibreOffice not found: no soffice on PATH and no LibreOffice.app in /Applications or ~/Applications",
                "hint": "brew install --cask libreoffice",
            },
            code=1,
        )

    try:
        with tempfile.TemporaryDirectory(prefix="spreadsheet-recalc-lo-") as tmp:
            profile_root = Path(tmp) / "profile"
            ensure_macro(soffice, profile_root)
            result = recalculate(soffice, profile_root, workbook, timeout)
    except subprocess.TimeoutExpired:
        report(
            {
                "status": "error",
                "error": f"LibreOffice timed out after {timeout}s",
                "hint": f"rerun with a larger timeout: uv run scripts/recalc.py {workbook.name} {timeout * 3}",
            },
            code=1,
        )

    if result.returncode != 0:
        report(
            {
                "status": "error",
                "error": f"LibreOffice exited with status {result.returncode}",
                "stdout": result.stdout.strip(),
                "stderr": result.stderr.strip(),
            },
            code=1,
        )

    try:
        audit_result = audit(workbook)
    except Exception as exc:
        report({"status": "error", "error": f"could not audit workbook: {exc}"}, code=1)

    if audit_result["status"] == "recalc_incomplete":
        audit_result["hint"] = "no cached values were written; close other LibreOffice instances and rerun"

    exit_code = 0 if audit_result["status"] == "success" or args.soft else 1
    report(audit_result, code=exit_code)


if __name__ == "__main__":
    main()
