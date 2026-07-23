#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.12"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Smoke tests for the spreadsheet skill helper scripts."""

from __future__ import annotations

import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent
PEEK = ROOT / "peek.py"
PROFILE = ROOT / "profile.py"
RECALC = ROOT / "recalc.py"


def run(args: list[str], check: bool = True) -> subprocess.CompletedProcess[str]:
    result = subprocess.run(args, capture_output=True, text=True, check=False)
    if check and result.returncode != 0:
        raise AssertionError(
            f"command failed: {' '.join(args)}\nstdout:\n{result.stdout}\nstderr:\n{result.stderr}"
        )
    return result


def run_json(args: list[str], check: bool = True) -> tuple[subprocess.CompletedProcess[str], dict]:
    result = run(args, check=check)
    return result, json.loads(result.stdout)


def test_peek(tmp: Path) -> None:
    good = tmp / "good.tsv"
    good.write_text("asset\tamount\tnote\nETH\t1.0\tok\nBTC\t-\tsafe\n", encoding="utf-8")

    _, auto = run_json([sys.executable, str(PEEK), str(good), "--strict", "--house"])
    _, python = run_json([sys.executable, str(PEEK), str(good), "--engine", "python", "--strict"])
    assert auto["status"] == "ok"
    assert auto["engine"] == "python-fast-unquoted"
    assert auto["data_rows"] == python["data_rows"] == 2
    assert auto["dash_null_cells"] == python["dash_null_cells"] == 1
    assert auto["header"] == python["header"]

    _, redacted = run_json([sys.executable, str(PEEK), str(good), "--redact-samples"])
    assert redacted["sample"][0] == ["<redacted>", "<redacted>", "<redacted>"]
    assert redacted["sample"][1][1] == "-"

    ragged = tmp / "ragged.tsv"
    ragged.write_text("asset\tamount\nETH\t1\nBTC\n", encoding="utf-8")
    result, payload = run_json([sys.executable, str(PEEK), str(ragged), "--strict"], check=False)
    assert result.returncode == 1
    assert payload["status"] == "issues_found"
    assert any(item["code"] == "ragged_rows" for item in payload["issues"])


def test_profile(tmp: Path) -> None:
    risky = tmp / "risky.tsv"
    risky.write_text("asset\tamount\tnote\nETH\t1.0\t=cmd\nBTC\t-\t+cmd\n", encoding="utf-8")
    _, payload = run_json(
        [sys.executable, str(PROFILE), str(risky), "--redact-samples", "--top", "2"]
    )
    assert payload["schema_version"] == 2
    assert payload["status"] == "ok"
    assert payload["formula_prefix_cells"]["count"] == 2
    assert payload["formula_prefix_cells"]["first"][0]["value"] == "<redacted>"
    assert "recommendations" not in payload
    assert payload["stats"]["available"] in (True, False)

    _, external = run_json([sys.executable, str(PROFILE), str(risky), "--external-data"])
    assert external["status"] == "issues_found"

    high_cardinality = tmp / "high-cardinality.tsv"
    high_cardinality.write_text("value\nfirst\nsecond\n", encoding="utf-8")
    _, cardinality = run_json([sys.executable, str(PROFILE), str(high_cardinality)])
    assert "recommendations" not in cardinality

    workbook = tmp / "metadata.xlsx"
    wb = Workbook()
    wb.save(workbook)
    _, workbook_profile = run_json([sys.executable, str(PROFILE), str(workbook)])
    assert workbook_profile["schema_version"] == 2
    assert "recommendations" not in workbook_profile


def test_recalc(tmp: Path) -> None:
    if shutil.which("soffice") is None and not Path("/Applications/LibreOffice.app/Contents/MacOS/soffice").is_file():
        print("skip recalc: LibreOffice not found")
        return

    clean = tmp / "clean.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["A2"] = 2
    ws["A3"] = "=SUM(A1:A2)"
    wb.save(clean)
    result, payload = run_json([sys.executable, str(RECALC), str(clean), "30"])
    assert result.returncode == 0
    assert payload["status"] == "success"

    bad = tmp / "bad.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 0
    ws["A2"] = "=1/A1"
    wb.save(bad)
    result, payload = run_json([sys.executable, str(RECALC), str(bad), "30"], check=False)
    assert result.returncode == 1
    assert payload["status"] == "errors_found"
    assert payload["total_errors"] == 1


def main() -> None:
    with tempfile.TemporaryDirectory(prefix="spreadsheet-skill-tests-") as tmp_dir:
        tmp = Path(tmp_dir)
        test_peek(tmp)
        test_profile(tmp)
        test_recalc(tmp)
    print("spreadsheet helper tests passed")


if __name__ == "__main__":
    main()
